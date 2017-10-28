#! /usr/bin/env/python

import argparse
from os.path import abspath
import shutil

from openpyxl import load_workbook
import pandas as pd
import win32com.client as win32
from xlrd import open_workbook
from xlutils.copy import copy


# PARSER = argparse.ArgumentParser(
#     description='Update the coordinated daily activity pattern calibration '
#     'workbooks and uec.')
# PARSER.add_argument(
#     'iteration', metavar='Iteration', type=int, default=0,
#     help='The current calibration iteration number.'
# )
# PARSER.add_argument(
#     'type', metavar='Type', type=str, default='AO', choices=['AO', 'CDAP'],
#     help='The type of update to perform.'
# )
# PARSER.add_argument(
#     '-i', '--input_path', metavar='Input_Path', type=str,
#     default='../../newpopsyn',
#     help='The path to the directory containing the abm output and uec ' +
#     'directories.')
# PARSER.add_argument(
#     '-o', '--output_path', metavar='Output_Path', type=str, default='.',
#     help='The path to the directory containing the model calibration files.')
# ARGS = PARSER.parse_args()


def replace_values(dest, data):
    """Replace the values in dest with those in data.

    Parameters
    ----------
    dest : tuple
        Tuple containing openpyxl.cell.cell's.
    data : array
        Array containing data to write into the cells in dest.
    """
    if len(dest) != len(data):
        raise ValueError('Length of dest and data should be the same.')
    for cell, value in zip(dest, data):
        cell.value = value


def exec_formulas(cal_path):
    """Use excel to execute formulas in workbook.

    Parameters
    ----------
    cal_path : string
        Path to a calibration file.

    """
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    workbook = excel.Workbooks.Open(abspath(cal_path))
    workbook.Save()
    workbook.Close()
    excel.Quit()


def update(iter_, input_path, output_path, method='AO'):
    """Aggregate model results, calculate constants, and update the UEC.

    Parameters
    ----------
    iter_ : int
        The calibration iteration number.
    input_path : str
        The relative path to the directory containing the output and uec
        directories.
    output_path : str
        The relative path to the directory containing the calibration files.
    method : str, 'AO' | 'CDAP'
        The type of update to perform.
        Default : 'AO'
        Valid Options :
        - 'AO' : Update AutoOwnership
        - 'CDAP' : Update CoordinatedDailyActivityPattern

    """
    files = {
        'AO': ['AutoOwnership', 'aoResults', '1_AO Calibration',
               update_ao],
        'CDAP': ['CoordinatedDailyActivityPattern', 'personData_3',
                 '2_CDAP Calibration', update_cdap]}

    cal_path = output_path + '/{}_{}.xlsx'.format(files[method][2], iter_)
    uec_path = input_path + '/uec/{}.xls'.format(files[method][0])
    shutil.copy2(input_path + '/output/{}.csv'.format(files[method][1]),
                 output_path + '/{}_{}.csv'.format(files[method][1], iter_))
    results = pd.read_csv(
        output_path + '/{}_{}.csv'.format(files[method][1], iter_))
    if iter_ < 1:
        wb_name = output_path + '/{}.xlsx'.format(files[method][2])
    else:
        wb_name = output_path + \
            '/{}_{}.xlsx'.format(files[method][2], iter_ - 1)
    files[method][3](iter_, wb_name, results, uec_path, cal_path)


def update_ao(iter_, wb_name, results, uec_path, cal_path):
    """Aggregate model results, calculate constants, and update the UEC.

    Parameters
    ----------
    iter_ : int
        The calibration iteration number.
    wb_name : str
        Path to the calibration workbook.
    results : pandas.DataFrame
        DataFrame of the results generated by the model.
    uec_path : str
        Path to the uec file.
    cal_path : str
        Path to the calibration file.

    """
    if iter_ > 0:
        prev_const = read_values(wb_name, 3, 11, 5, sheet_num=0)

    workbook = load_workbook(wb_name)
    replace_values(workbook['_data']['B'][1:6],
                   results.groupby('AO').size().values)

    if iter_ > 0:
        replace_values(workbook['AO']['K'][3:8], prev_const)
    else:
        prev_const = read_values(uec_path, 81, 6, 5, axis=1)
        replace_values(workbook['AO']['K'][3:8], prev_const)

    workbook.save(cal_path)
    workbook.close()

    exec_formulas(cal_path)

    workbook = load_workbook(cal_path, data_only=True)
    new_constants = [cell.value for cell in workbook['AO']['L'][3:8]]
    workbook.close()

    update_uec(uec_path, 81, 6, new_constants, axis=1)


def update_uec(uec_path, startx, starty, values, axis=0, sheet_num=1):
    """Update a uec file column starting at the given cell with the new values.

    Parameters
    ----------
    uec_path : string
        Path to the uec file.
    startx : int
        Starting cell row index.
    starty : int
        Starting cell column index.
    values : array-like
        Values to update.
    axis : int, default : 0
        The axis to update values along. Rows = 0, columns = 1.
    sheet_num : int, default : 1
        The sheet number to use.

    """
    uec = open_workbook(uec_path, formatting_info=True)
    workbook = copy(uec)
    sheet = workbook.get_sheet(sheet_num)
    if axis == 0:
        for idx, val in enumerate(values):
            sheet.write(startx + idx, starty, val)
    else:
        for idx, val in enumerate(values):
            sheet.write(startx, starty + idx, val)
    workbook.save(uec_path)
    uec.release_resources()


def read_values(filename, startx, starty, length, axis=0, sheet_num=1):
    """Read data from a uec file starting at the given cell.

    Parameters
    ----------
    filename : string
        Path to the uec file.
    startx : int
        Starting cell row index.
    starty : int
        Starting cell column index.
    length : int
        The number of values to read.
    axis : int, default : 0
        The axis to update values along. Rows = 0, columns = 1.
    sheet_num : int, default : 1
        The sheet number to use.

    Returns
    -------
    vals : array-like
        Array of constants read from uec file.

    """
    uec = open_workbook(filename)
    sheet = uec.sheet_by_index(sheet_num)
    if axis == 0:
        vals = [sheet.cell_value(rowx=startx + idx, colx=starty)
                for idx in range(length)]
    else:
        vals = [sheet.cell_value(rowx=startx, colx=starty + idx)
                for idx in range(length)]
    uec.release_resources()
    return vals


def update_cdap(iter_, wb_name, results, uec_path, cal_path):
    """Aggregate model results, calculate constants, and update the UEC.

    Parameters
    ----------
    iter_ : int
        The calibration iteration number.
    wb_name : str
        Path to the calibration workbook.
    results : pandas.DataFrame
        DataFrame of the results generated by the model.
    uec_path : str
        Path to the uec file.
    cal_path : str
        Path to the calibration file.

    """
    res = results.groupby(['type', 'activity_pattern']).size()\
        .reset_index()
    names = {'Child too young for school': 'Pre-school',
             'Non-worker': 'Non-working Adult',
             'Retired': 'Non-working Senior',
             'Student of driving age': 'Driving Age Student',
             'Student of non-driving age': 'Non-driving Student',
             'University Student': 'College Student'}
    for orig, final in names.items():
        res.loc[res.type == orig, 'type'] = final
    res_vals = res.sort_values(['type', 'activity_pattern'])[0].values

    if iter_ > 0:
        prev_m_const = read_values(wb_name, 30, 9, 8, sheet_num=0)
        prev_n_const = read_values(wb_name, 30, 10, 8, sheet_num=0)

    workbook = load_workbook(wb_name)
    replace_values(workbook['_data']['E'][1:23], res_vals)

    if iter_ > 0:
        replace_values(workbook['CDAP']['C'][29:37], prev_m_const)
        replace_values(workbook['CDAP']['D'][29:37], prev_n_const)
    else:
        prev_m_const = read_values(uec_path, 88, 6, 8)
        prev_n_const = read_values(uec_path, 88, 7, 8)
        replace_values(workbook['CDAP']['C'][29:37], prev_m_const)
        replace_values(workbook['CDAP']['D'][29:37], prev_n_const)

    workbook.save(cal_path)
    workbook.close()

    exec_formulas(cal_path)

    new_m_const = read_values(cal_path, 30, 9, 8, sheet_num=0)
    new_n_const = read_values(cal_path, 30, 10, 8, sheet_num=0)

    update_uec(uec_path, 88, 6, new_m_const)
    update_uec(uec_path, 88, 7, new_n_const)


if __name__ == '__main__':
    update()
