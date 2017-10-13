#! /usr/bin/env/python

import argparse
import os.path as osp
import shutil

from openpyxl import load_workbook
import pandas as pd
import win32com.client as win32
from xlrd import open_workbook
from xlutils.copy import copy


PARSER = argparse.ArgumentParser(
    description='Update the auto ownership workbooks and uec.')
PARSER.add_argument(
    'iteration', metavar='Iteration', type=int, default=0,
    help='The current calibration iteration number.'
)
PARSER.add_argument(
    '-i', '--input_path', metavar='Input_Path', type=str,
    default='../../newpopsyn',
    help='The path to the directory containing the abm output and uec ' +
    'directories.')
PARSER.add_argument(
    '-o', '--output_path', metavar='Output_Path', type=str, default='.',
    help='The path to the directory containing the model calibration files.')
ARGS = PARSER.parse_args()


def replace_values(dest, data):
    """Replace the values in dest with those in data.

    Parameters
    ----------
    dest : tuple
        Tuple containing openpyxl.cell.cell's.
    data : array
        Array containing data to write into the cells in dest.
    """
    if len(dest) != len(data):
        raise ValueError('Length of dest and data should be the same.')
    for cell, value in zip(dest, data):
        cell.value = value


def update_auto_ownership(iter_, input_path, output_path):
    """Aggregate model results, calculate constants, and update the UEC.

    Parameters
    ----------
    iter_ : int
        The calibration iteration number.
    input_path : str
        The relative path to the directory containing the output and uec
        directories.
    output_path : str
        The relative path to the directory containing the calibration files.

    """
    uec_path = input_path + '/uec/AutoOwnership.xls'
    shutil.copy2(input_path + '/output/aoResults.csv',
                 output_path + f'/aoResults-{iter_}.csv')
    model_results = pd.read_csv(output_path + f'/aoResults-{iter_}.csv')
    if iter_ <= 1:
        wb_name = output_path + '/1_AO Calibration.xlsx'
    else:
        wb_name = output_path + f'/1_AO Calibration_{iter_ - 1}.xlsx'

    workbook = load_workbook(wb_name, data_only=True)
    prev_constants = [cell.value for cell in workbook['AO']['L'][3:8]]
    workbook.close()

    workbook = load_workbook(wb_name)
    replace_values(workbook['_data']['B'][1:6],
                   model_results.groupby('AO').HHID.count().values)
    if iter_ > 0:
        cal_out = output_path + f'/1_AO Calibration_{iter_}.xlsx'
        replace_values(workbook['AO']['K'][3:8], prev_constants)
    else:
        cal_out = output_path + '/1_AO Calibration.xlsx'
        ao_uec = open_workbook(uec_path)
        ao_sheet = ao_uec.sheet_by_index(1)
        prev_constants = [
            ao_sheet.cell_value(rowx=81, colx=6 + idx) for idx in range(5)]
        ao_uec.release_resources()
        replace_values(workbook['AO']['K'][3:8], prev_constants)

    workbook.save(cal_out)
    workbook.close()

    excel = win32.gencache.EnsureDispatch('Excel.Application')
    workbook = excel.Workbooks.Open(osp.abspath(cal_out))
    workbook.Save()
    workbook.Close()
    excel.Quit()

    workbook = load_workbook(cal_out, data_only=True)
    new_constants = [cell.value for cell in workbook['AO']['L'][3:8]]
    workbook.close()

    ao_uec = open_workbook(uec_path, formatting_info=True)
    workbook = copy(ao_uec)
    auto_ownership = workbook.get_sheet(1)
    for idx, val in enumerate(new_constants):
        auto_ownership.write(81, 6 + idx, val)
    workbook.save(uec_path)
    ao_uec.release_resources()


if __name__ == '__main__':
    update_auto_ownership(
        ARGS.iteration, ARGS.input_path, ARGS.output_path)
