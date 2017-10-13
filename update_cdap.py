#! /usr/bin/env/python

import argparse
from os.path import abspath
import shutil

from openpyxl import load_workbook
import pandas as pd
import win32com.client as win32
from xlrd import open_workbook
from xlutils.copy import copy


PARSER = argparse.ArgumentParser(
    description='Update the coordinated daily activity pattern calibration '
    'workbooks and uec.')
PARSER.add_argument(
    'iteration', metavar='Iteration', type=int, default=0,
    help='The current calibration iteration number.'
)
PARSER.add_argument(
    '-i', '--input_path', metavar='Input_Path', type=str,
    default='../../newpopsyn',
    help='The path to the directory containing the abm output and uec ' +
    'directories.')
PARSER.add_argument(
    '-o', '--output_path', metavar='Output_Path', type=str, default='.',
    help='The path to the directory containing the model calibration files.')
ARGS = PARSER.parse_args()


def replace_values(dest, data):
    """Replace the values in dest with those in data.

    Parameters
    ----------
    dest : tuple
        Tuple containing openpyxl.cell.cell's.
    data : array
        Array containing data to write into the cells in dest.
    """
    if len(dest) != len(data):
        raise ValueError('Length of dest and data should be the same.')
    for cell, value in zip(dest, data):
        cell.value = value


def update_cdap(iter_, input_path, output_path):
    """Aggregate model results, calculate constants, and update the UEC.

    Parameters
    ----------
    iter_ : int
        The calibration iteration number.
    input_path : str
        The relative path to the directory containing the output and uec
        directories.
    output_path : str
        The relative path to the directory containing the calibration files.

    """
    uec_path = input_path + '/uec/CoordinatedDailyActivityPattern.xls'
    shutil.copy2(input_path + '/output/personData_3.csv',
                 output_path + f'/personData_{iter_}.csv')

    results = pd.read_csv(output_path + f'/personData_{iter_}.csv')
    res_vals = results.groupby(['type', 'activity_pattern']).size()\
        .reset_index()
    res_vals.loc[res_vals.type == 'Child too young for school', 'type'] = \
        'Pre-school'
    res_vals.loc[res_vals.type == 'Non-worker', 'type'] = 'Non-working Adult'
    res_vals.loc[res_vals.type == 'Retired', 'type'] = 'Non-working Senior'
    res_vals.loc[res_vals.type == 'Student of driving age', 'type'] = \
        'Driving Age Student'
    res_vals.loc[res_vals.type == 'Student of non-driving age', 'type'] = \
        'Non-driving Student'
    res_vals.loc[res_vals.type == 'University Student', 'type'] = \
        'College Student'
    res_vals = res_vals.sort_values(['type', 'activity_pattern'])[0].values

    if iter_ <= 1:
        wb_name = output_path + '/2_CDAP Calibration.xlsx'
    else:
        wb_name = output_path + f'/2_CDAP Calibration_{iter_ - 1}.xlsx'

    if iter_ > 0:
        workbook = load_workbook(wb_name, data_only=True)
        prev_m_const = [workbook['CDAP'].cell(row=30 + idx, column=9).value
                        for idx in range(8)]
        prev_n_const = [workbook['CDAP'].cell(row=30 + idx, column=10).value
                        for idx in range(8)]
        workbook.close()

    workbook = load_workbook(wb_name)
    replace_values(workbook['_data']['E'][1:23], res_vals)

    if iter_ > 0:
        cal_out = output_path + f'/2_CDAP Calibration_{iter_}.xlsx'
        replace_values(workbook['CDAP']['C'][29:37], prev_m_const)
        replace_values(workbook['CDAP']['D'][29:37], prev_n_const)
    else:
        cal_out = output_path + '/2_CDAP Calibration.xlsx'
        uec = open_workbook(uec_path)
        sheet = uec.sheet_by_index(1)
        prev_m_const = [
            sheet.cell_value(rowx=88 + idx, colx=6) for idx in range(8)]
        prev_n_const = [
            sheet.cell_value(rowx=88 + idx, colx=7) for idx in range(8)]
        uec.release_resources()
        replace_values(workbook['CDAP']['C'][29:37], prev_m_const)
        replace_values(workbook['CDAP']['D'][29:37], prev_n_const)

    workbook.save(cal_out)
    workbook.close()

    excel = win32.gencache.EnsureDispatch('Excel.Application')
    workbook = excel.Workbooks.Open(abspath(cal_out))
    workbook.Save()
    workbook.Close()
    excel.Quit()

    workbook = load_workbook(cal_out, data_only=True)
    new_m_const = [workbook['CDAP'].cell(row=30 + idx, column=9).value
                   for idx in range(8)]
    new_n_const = [workbook['CDAP'].cell(row=30 + idx, column=10).value
                   for idx in range(8)]
    workbook.close()

    cdap_uec = open_workbook(uec_path, formatting_info=True)
    workbook = copy(cdap_uec)
    cdap = workbook.get_sheet(1)
    for idx, val in enumerate(new_m_const):
        cdap.write(88 + idx, 6, val)
    for idx, val in enumerate(new_n_const):
        cdap.write(88 + idx, 7, val)
    workbook.save(uec_path)
    cdap_uec.release_resources()


if __name__ == '__main__':
    update_cdap(
        ARGS.iteration, ARGS.input_path, ARGS.output_path)
